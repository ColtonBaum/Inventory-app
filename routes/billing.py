# routes/billing.py
from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, session, current_app, make_response, abort
)
from models import ItemPrice, Trailer, InventoryResponse, WarehouseProduct, WarehouseOrder, WarehouseOrderLine, SpecialtyTool
from database import db
from utils.tooling_lists import get_tooling_list
from functools import wraps
from collections import defaultdict
from datetime import datetime
import io

billing_bp = Blueprint('billing', __name__, url_prefix='/billing')


# ---------- Auth ----------
def billing_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('billing_auth'):
            return redirect(url_for('billing.login', next=request.url))
        return f(*args, **kwargs)
    return decorated


@billing_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        pw = request.form.get('password', '')
        if pw == current_app.config.get('BILLING_PASSWORD', ''):
            session['billing_auth'] = True
            next_url = request.args.get('next') or url_for('billing.billing_dashboard')
            return redirect(next_url)
        flash('Incorrect password.', 'danger')
    return render_template('billing_login.html')


@billing_bp.route('/logout')
def logout():
    session.pop('billing_auth', None)
    flash('Logged out of billing.', 'info')
    return redirect(url_for('inventory.dashboard'))


# ---------- Dashboard ----------
@billing_bp.route('/')
@billing_required
def billing_dashboard():
    q = (request.args.get('q') or '').strip().lower()
    trailers = Trailer.query.filter(Trailer.status == 'Completed').order_by(Trailer.id.desc()).all()
    if q:
        trailers = [t for t in trailers if q in (t.job_name or '').lower()
                     or q in (t.job_number or '').lower()
                     or q in str(t.id)]
    return render_template('billing_dashboard.html', trailers=trailers, q=q)


# ---------- Pricing Management ----------

# ---------- Shared helper: compute invoice line items from live DB ----------
def _compute_line_items(trailer):
    """Return (line_items, total) using current ItemPrice values."""
    import json as _json
    responses = InventoryResponse.query.filter_by(trailer_id=trailer.id).all()
    list_name = (trailer.tooling_list_name or trailer.inventory_type or '').strip()
    tooling_list = get_tooling_list(list_name) or []

    expected_map = {}
    for item in tooling_list:
        num = item.get('Item Number', '').strip()
        expected_map[num] = {
            'item_name': item.get('Item Name', ''),
            'quantity': int(item.get('Quantity', 0)) if str(item.get('Quantity', 0)).isdigit() else 0,
        }

    response_map = defaultdict(lambda: {'missing': 0, 'redtag': 0, 'note': ''})
    for r in responses:
        key = r.item_number
        if r.status == 'Missing':
            response_map[key]['missing'] += (r.quantity or 0)
            if r.note:
                response_map[key]['note'] = r.note
        elif r.status == 'Red Tag':
            response_map[key]['redtag'] += (r.quantity or 0)
            if r.note:
                response_map[key]['note'] = r.note

    price_map = {p.item_number: p.price for p in ItemPrice.query.all()}

    line_items = []
    total = 0.0

    for num, info in expected_map.items():
        resp = response_map.get(num, {})
        missing_qty = resp.get('missing', 0)
        redtag_qty = resp.get('redtag', 0)
        billable_qty = missing_qty + redtag_qty
        if billable_qty <= 0:
            continue
        unit_price = price_map.get(num, 0.0)
        line_total = unit_price * billable_qty
        total += line_total
        line_items.append({
            'item_number': num,
            'item_name': info['item_name'],
            'missing_qty': missing_qty,
            'redtag_qty': redtag_qty,
            'billable_qty': billable_qty,
            'unit_price': unit_price,
            'line_total': line_total,
            'note': resp.get('note', ''),
        })

    # Extra tooling
    extra_map = defaultdict(lambda: {'missing': 0, 'redtag': 0, 'item_name': '', 'note': ''})
    for r in responses:
        if (r.category or '').strip().lower() == 'extra tooling':
            key = r.item_number
            extra_map[key]['item_name'] = r.item_name
            if r.status == 'Missing':
                extra_map[key]['missing'] += (r.quantity or 0)
                if r.note:
                    extra_map[key]['note'] = r.note
            elif r.status == 'Red Tag':
                extra_map[key]['redtag'] += (r.quantity or 0)
                if r.note:
                    extra_map[key]['note'] = r.note

    for num, info in extra_map.items():
        billable_qty = info['missing'] + info['redtag']
        if billable_qty <= 0:
            continue
        unit_price = price_map.get(num, 0.0)
        line_total = unit_price * billable_qty
        total += line_total
        line_items.append({
            'item_number': num,
            'item_name': info['item_name'],
            'missing_qty': info['missing'],
            'redtag_qty': info['redtag'],
            'billable_qty': billable_qty,
            'unit_price': unit_price,
            'line_total': line_total,
            'note': info.get('note', ''),
        })

    line_items.sort(key=lambda x: (x['item_name'] or '').lower())
    return line_items, total


# ---------- Generate Billing Invoice for a Trailer ----------
@billing_bp.route('/invoice/<int:trailer_id>')
@billing_required
def generate_billing_invoice(trailer_id):
    import json as _json
    from models import Invoice as _Invoice
    trailer = Trailer.query.get_or_404(trailer_id)
    invoice = _Invoice.query.filter_by(trailer_id=trailer_id).order_by(_Invoice.id.desc()).first()

    # Use frozen snapshot if invoice has been marked billed
    if invoice and invoice.billed and invoice.line_items_json:
        line_items = _json.loads(invoice.line_items_json)
        total = sum(li['line_total'] for li in line_items)
    else:
        line_items, total = _compute_line_items(trailer)

    return render_template(
        'billing_invoice.html',
        trailer=trailer,
        line_items=line_items,
        total=total,
        now=datetime.now,
    )


# ---------- Warehouse Stock ----------
@billing_bp.route('/warehouse')
@billing_required
def warehouse_inventory():
    q = (request.args.get('q') or '').strip().lower()
    products = WarehouseProduct.query.order_by(WarehouseProduct.item_name).all()
    if q:
        products = [p for p in products if q in (p.item_name or '').lower() or q in (p.item_number or '').lower()]
    low_stock = [p for p in products if p.quantity_on_hand <= p.reorder_point]
    return render_template('billing_inventory.html', products=products, low_stock=low_stock, q=q)


@billing_bp.route('/warehouse/product/<int:product_id>/edit', methods=['GET', 'POST'])
@billing_required
def edit_product(product_id):
    product = WarehouseProduct.query.get_or_404(product_id)
    if request.method == 'POST':
        product.item_name = request.form.get('item_name', product.item_name)
        product.quantity_on_hand = int(request.form.get('quantity_on_hand', 0) or 0)
        product.reorder_point = int(request.form.get('reorder_point', 0) or 0)
        try:
            product.unit_cost = float(request.form.get('unit_cost', 0) or 0)
        except ValueError:
            product.unit_cost = 0.0
        db.session.commit()
        flash('Product updated.', 'success')
        return redirect(url_for('billing.warehouse_inventory'))
    return render_template('billing_edit_product.html', product=product)


@billing_bp.route('/warehouse/product/add', methods=['POST'])
@billing_required
def add_product():
    item_number = (request.form.get('item_number') or '').strip()
    if not item_number:
        flash('Item number is required.', 'danger')
        return redirect(url_for('billing.warehouse_inventory'))
    existing = WarehouseProduct.query.filter_by(item_number=item_number).first()
    if existing:
        flash('A product with that item number already exists.', 'warning')
        return redirect(url_for('billing.warehouse_inventory'))
    p = WarehouseProduct(
        item_number=item_number,
        item_name=(request.form.get('item_name') or '').strip(),
        quantity_on_hand=int(request.form.get('quantity_on_hand', 0) or 0),
        reorder_point=int(request.form.get('reorder_point', 0) or 0),
        unit_cost=float(request.form.get('unit_cost', 0) or 0),
    )
    db.session.add(p)
    db.session.commit()
    flash('Product added.', 'success')
    return redirect(url_for('billing.warehouse_inventory'))


# ---------- Warehouse Orders ----------
@billing_bp.route('/warehouse/orders')
@billing_required
def warehouse_orders():
    status_filter = request.args.get('status', '')
    q = WarehouseOrder.query.order_by(WarehouseOrder.created_at.desc())
    if status_filter:
        q = q.filter(WarehouseOrder.status == status_filter)
    orders = q.all()
    trailers = {t.id: t for t in Trailer.query.all()}
    return render_template('billing_orders.html', orders=orders, trailers=trailers, status_filter=status_filter)


@billing_bp.route('/warehouse/orders/new', methods=['GET', 'POST'])
@billing_required
def new_order():
    if request.method == 'POST':
        trailer_id_raw = request.form.get('trailer_id') or None
        trailer_id = int(trailer_id_raw) if trailer_id_raw else None
        notes = (request.form.get('notes') or '').strip()
        order = WarehouseOrder(trailer_id=trailer_id, status='Pending', notes=notes)
        db.session.add(order)
        db.session.flush()  # get order.id

        item_numbers = request.form.getlist('item_number')
        item_names = request.form.getlist('item_name')
        quantities = request.form.getlist('quantity')
        for num, name, qty_raw in zip(item_numbers, item_names, quantities):
            num = num.strip()
            if not num:
                continue
            try:
                qty = int(qty_raw or 0)
            except ValueError:
                qty = 0
            if qty <= 0:
                continue
            line = WarehouseOrderLine(order_id=order.id, item_number=num, item_name=name.strip(), quantity=qty)
            db.session.add(line)
        db.session.commit()
        flash('Order created.', 'success')
        return redirect(url_for('billing.view_order', order_id=order.id))

    trailers = Trailer.query.order_by(Trailer.id.desc()).all()
    products = WarehouseProduct.query.order_by(WarehouseProduct.item_name).all()
    return render_template('billing_order_new.html', trailers=trailers, products=products)


@billing_bp.route('/warehouse/orders/<int:order_id>')
@billing_required
def view_order(order_id):
    order = WarehouseOrder.query.get_or_404(order_id)
    trailer = Trailer.query.get(order.trailer_id) if order.trailer_id else None
    return render_template('billing_order_view.html', order=order, trailer=trailer)


@billing_bp.route('/warehouse/orders/<int:order_id>/status', methods=['POST'])
@billing_required
def update_order_status(order_id):
    order = WarehouseOrder.query.get_or_404(order_id)
    new_status = request.form.get('status', order.status)
    if new_status in ('Pending', 'Cancelled'):
        order.status = new_status
        db.session.commit()
        flash(f'Order status updated to {new_status}.', 'success')
    return redirect(url_for('billing.view_order', order_id=order_id))


@billing_bp.route('/warehouse/orders/<int:order_id>/bill', methods=['POST'])
@billing_required
def mark_order_billed(order_id):
    """Mark an order as billed: match items by name, deduct inventory, snapshot prices."""
    order = WarehouseOrder.query.get_or_404(order_id)
    if order.billed:
        flash('Order is already billed.', 'warning')
        return redirect(url_for('billing.view_order', order_id=order_id))

    # Build lookup maps: name (lower) -> product, item_number (upper) -> price
    products_by_name = {(p.item_name or '').strip().lower(): p
                        for p in WarehouseProduct.query.all() if p.item_name}
    price_map = {p.item_number.upper(): p.price for p in ItemPrice.query.all()}

    order_total = 0.0
    unmatched = []

    is_purchase = (order.order_type == 'PURCHASE')

    for line in order.lines:
        name_key = (line.item_name or '').strip().lower()
        product = products_by_name.get(name_key)

        if product:
            # SALE: remove from stock. PURCHASE: add to stock.
            if is_purchase:
                product.quantity_on_hand += line.quantity
            else:
                product.quantity_on_hand -= line.quantity
            unit_price = price_map.get(product.item_number.upper(), product.unit_cost or 0.0)
            line.unit_price = unit_price
            line.line_total = unit_price * line.quantity
            line.item_number = product.item_number
        else:
            line.unit_price = 0.0
            line.line_total = 0.0
            unmatched.append(line.item_name or '(unnamed)')

        order_total += line.line_total

    order.billed = True
    order.status = 'Billed'
    order.order_total = order_total
    db.session.commit()

    action = 'received into' if is_purchase else 'deducted from'
    if unmatched:
        flash(f'Order billed. {len(unmatched)} item(s) not matched in warehouse stock '
              f'(inventory not updated): {", ".join(unmatched[:5])}.', 'warning')
    else:
        flash(f'Order billed. Total: ${order_total:,.2f}. Inventory {action} stock.', 'success')
    return redirect(url_for('billing.view_order', order_id=order_id))


# ---------- Tooling List Management ----------
@billing_bp.route('/tooling-lists')
@billing_required
def tooling_lists_index():
    from models import ToolingListItem
    from sqlalchemy import func
    # Get all unique list names with item counts
    counts = (db.session.query(ToolingListItem.list_name, func.count(ToolingListItem.id))
              .group_by(ToolingListItem.list_name)
              .order_by(ToolingListItem.list_name)
              .all())
    return render_template('billing_tooling_lists.html', list_counts=counts)


@billing_bp.route('/tooling-lists/<list_name>')
@billing_required
def tooling_list_detail(list_name):
    from models import ToolingListItem
    items = (ToolingListItem.query.filter_by(list_name=list_name)
             .order_by(ToolingListItem.sort_order, ToolingListItem.id).all())
    return render_template('billing_tooling_list_detail.html', list_name=list_name, items=items)


@billing_bp.route('/tooling-lists/<list_name>/add', methods=['POST'])
@billing_required
def tooling_list_add_item(list_name):
    from models import ToolingListItem
    item_number = (request.form.get('item_number') or '').strip()
    item_name = (request.form.get('item_name') or '').strip()
    category = (request.form.get('category') or 'General').strip()
    try:
        quantity = int(request.form.get('quantity') or 0)
    except ValueError:
        quantity = 0
    # Put new items at the end
    max_sort = db.session.query(db.func.max(ToolingListItem.sort_order)).filter_by(list_name=list_name).scalar() or 0
    db.session.add(ToolingListItem(list_name=list_name, item_number=item_number,
                                   item_name=item_name, category=category,
                                   quantity=quantity, sort_order=max_sort+1))
    db.session.commit()
    flash(f'Item added to {list_name}.', 'success')
    return redirect(url_for('billing.tooling_list_detail', list_name=list_name))


@billing_bp.route('/tooling-lists/item/<int:item_id>/edit', methods=['POST'])
@billing_required
def tooling_list_edit_item(item_id):
    from models import ToolingListItem
    item = ToolingListItem.query.get_or_404(item_id)
    item.item_number = (request.form.get('item_number') or item.item_number).strip()
    item.item_name = (request.form.get('item_name') or item.item_name).strip()
    item.category = (request.form.get('category') or item.category).strip()
    try:
        item.quantity = int(request.form.get('quantity') or item.quantity)
    except ValueError:
        pass
    db.session.commit()
    flash('Item updated.', 'success')
    return redirect(url_for('billing.tooling_list_detail', list_name=item.list_name))


@billing_bp.route('/tooling-lists/item/<int:item_id>/delete', methods=['POST'])
@billing_required
def tooling_list_delete_item(item_id):
    from models import ToolingListItem
    item = ToolingListItem.query.get_or_404(item_id)
    list_name = item.list_name
    db.session.delete(item)
    db.session.commit()
    flash('Item removed.', 'info')
    return redirect(url_for('billing.tooling_list_detail', list_name=list_name))


@billing_bp.route('/tooling-lists/new-list', methods=['POST'])
@billing_required
def tooling_list_create():
    list_name = (request.form.get('list_name') or '').strip()
    if not list_name:
        flash('List name is required.', 'danger')
        return redirect(url_for('billing.tooling_lists_index'))
    flash(f'List "{list_name}" created. Add items to it now.', 'success')
    return redirect(url_for('billing.tooling_list_detail', list_name=list_name))


# ---------- Warehouse Excel Import ----------
@billing_bp.route('/warehouse/import', methods=['GET', 'POST'])
@billing_required
def import_warehouse():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename.endswith('.xlsx'):
            flash('Please upload a .xlsx file.', 'danger')
            return redirect(url_for('billing.import_warehouse'))

        try:
            import openpyxl
            # read_only=True streams the file instead of loading everything into memory
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        except Exception as e:
            flash(f'Could not read Excel file: {e}', 'danger')
            return redirect(url_for('billing.import_warehouse'))

        # Find the inventory/products sheet (prefer named sheets over active)
        PRODUCT_SHEET_NAMES = {'products', 'product', 'inventory', 'stock', 'warehouse stock',
                                'items', 'catalog', 'warehouse', 'item list'}
        ws = None
        for sname in wb.sheetnames:
            if sname.strip().lower() in PRODUCT_SHEET_NAMES:
                ws = wb[sname]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]  # fall back to first sheet

        # Scan up to 10 rows to find the actual header row
        ITEM_NUM_CANDIDATES = {
            'item number', 'item_number', 'item #', 'item no', 'part number', 'part #',
            'part no', 'sku', 'product number', 'product #', 'number', 'no', 'id',
            'item id', 'product id', 'product_id', 'code', 'item code', 'part code',
        }
        header_row_idx = None
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            row_vals = [(str(c or '').strip().lower()) for c in row]
            if any(v in ITEM_NUM_CANDIDATES for v in row_vals):
                header_row_idx = row_idx
                headers = row_vals
                break

        def find_col(candidates):
            for c in candidates:
                if c in headers:
                    return headers.index(c)
            return None

        col_num = find_col([
            'item_number', 'item number', 'item #', 'item no', 'part number', 'part #',
            'part no', 'product_id', 'product id', 'sku', 'product number', 'product #',
            'number', 'no', 'id', 'item id', 'code', 'item code', 'part code',
        ])
        col_name = find_col([
            'name', 'item name', 'item_name', 'product name', 'description', 'desc',
            'item description', 'product description', 'title',
        ])
        col_qty = find_col([
            'inventory_on_hand', 'inventory on hand', 'quantity', 'qty', 'on hand',
            'quantity on hand', 'qty on hand', 'stock', 'stock on hand',
            'inventory', 'count', 'available', 'balance',
        ])
        col_reorder = find_col([
            'reorder_point', 'reorder point', 'reorder', 'reorder qty', 'min stock',
            'minimum', 'min qty', 'min', 'reorder level', 'minimum stock', 'min quantity',
        ])
        col_price = find_col([
            'unit cost', 'unit_cost', 'price', 'unit price', 'cost', 'rate', 'each',
            'purchase_price', 'purchase price', 'unit cost ($)', 'price ($)',
            'cost ($)', 'sell price', 'list price', 'sales_price', 'sale price',
        ])

        if col_num is None or header_row_idx is None:
            sample = []
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                for cell in row:
                    v = str(cell or '').strip()
                    if v and v not in sample:
                        sample.append(v)
                if len(sample) > 20:
                    break
            flash(
                f'Could not find an "Item Number" column in the first 10 rows. '
                f'Values found: {", ".join(sample[:20]) or "(empty sheet)"}. '
                f'Rename your item number column to "Item Number", "ITEM_NUMBER", or "PRODUCT_ID".',
                'danger'
            )
            return redirect(url_for('billing.import_warehouse'))

        # --- Read PRICES sheet (PRODUCT_ID, EFFECTIVE_FROM_DATE, PURCHASE_PRICE, SALES_PRICE) ---
        # Pick the most recent price record with EFFECTIVE_FROM_DATE <= today per product
        from datetime import date as _date
        today = datetime.now().date()

        def _try_parse_date(val):
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, _date):
                return val
            s = str(val or '').strip()
            for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%m-%d-%Y',
                        '%B %d, %Y', '%b %d, %Y', '%d-%b-%Y', '%d/%m/%Y'):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
            return None

        PRICE_SHEET_NAMES = {'prices', 'price', 'pricing', 'price list', 'price book',
                              'rate sheet', 'pricelist'}
        price_from_sheet = {}  # item_number (upper) -> best price float
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in PRICE_SHEET_NAMES:
                pws = wb[sheet_name]
                # Find header row
                p_hdr_idx = None
                p_hdrs = []
                for r_idx, row in enumerate(pws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                    row_vals = [str(c or '').strip().lower() for c in row]
                    if any(v in ITEM_NUM_CANDIDATES for v in row_vals):
                        p_hdr_idx = r_idx
                        p_hdrs = row_vals
                        break
                if p_hdr_idx is None:
                    break

                def pfind(cands):
                    for c in cands:
                        if c in p_hdrs:
                            return p_hdrs.index(c)
                    return None

                p_col_id = pfind(['product_id', 'product id', 'item_number', 'item number',
                                   'item #', 'item no', 'sku', 'id', 'code'])
                p_col_date = pfind(['effective_from_date', 'effective from date', 'effective date',
                                     'date', 'price date', 'from date', 'start date'])
                p_col_purchase = pfind(['purchase_price', 'purchase price', 'cost', 'unit cost',
                                         'buy price', 'cost price'])
                p_col_sales = pfind(['sales_price', 'sale price', 'sales price', 'sell price',
                                      'price', 'unit price', 'list price'])

                if p_col_id is None:
                    break

                # Collect all rows: {item_upper -> [(date, purchase_price, sales_price)]}
                from collections import defaultdict as _dd
                price_records = _dd(list)
                for row in pws.iter_rows(min_row=p_hdr_idx + 1, values_only=True):
                    item_id = str(row[p_col_id] or '').strip().upper()
                    if not item_id:
                        continue
                    d = _try_parse_date(row[p_col_date]) if p_col_date is not None else None
                    try:
                        purchase = float(row[p_col_purchase] or 0) if p_col_purchase is not None else None
                    except (ValueError, TypeError):
                        purchase = None
                    try:
                        sales = float(row[p_col_sales] or 0) if p_col_sales is not None else None
                    except (ValueError, TypeError):
                        sales = None
                    price_records[item_id].append((d, purchase, sales))

                # For each product, pick the record with the most recent date <= today
                for item_id, records in price_records.items():
                    past = [(d, pu, sa) for d, pu, sa in records if d is not None and d <= today]
                    if past:
                        best = max(past, key=lambda x: x[0])
                    elif records:
                        best = min(
                            [(d, pu, sa) for d, pu, sa in records if d is not None],
                            key=lambda x: x[0],
                            default=records[0]
                        )
                    else:
                        continue
                    _, purchase_price, sales_price = best
                    # Use sales_price for billing invoices; fall back to purchase_price
                    price_from_sheet[item_id] = sales_price if sales_price else purchase_price
                break  # only process first matching sheet

        # Pre-load existing records with UPPERCASE keys for case-insensitive matching
        # This avoids duplicates without touching existing DB records
        existing_products = {p.item_number.upper(): p for p in WarehouseProduct.query.all()}
        existing_prices = {p.item_number.upper(): p for p in ItemPrice.query.all()}

        added = updated = priced = 0
        new_products = []
        new_prices = []

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            item_number = str(row[col_num] or '').strip().upper()  # normalize to uppercase
            if not item_number:
                continue
            item_name = str(row[col_name] or '').strip() if col_name is not None else ''
            try:
                qty = int(float(row[col_qty] or 0)) if col_qty is not None else None
            except (ValueError, TypeError):
                qty = None
            try:
                reorder = int(float(row[col_reorder] or 0)) if col_reorder is not None else None
            except (ValueError, TypeError):
                reorder = None
            # Price: from PRICES sheet first, then inline column
            price = price_from_sheet.get(item_number)
            if price is None and col_price is not None:
                try:
                    price = float(row[col_price] or 0) or None
                except (ValueError, TypeError):
                    price = None

            # Upsert WarehouseProduct
            if qty is not None or reorder is not None or item_name:
                wp = existing_products.get(item_number)
                if wp:
                    if item_name:
                        wp.item_name = item_name
                    if qty is not None:
                        wp.quantity_on_hand = qty
                    if reorder is not None:
                        wp.reorder_point = reorder
                    if price is not None:
                        wp.unit_cost = price
                    updated += 1
                else:
                    wp = WarehouseProduct(
                        item_number=item_number,
                        item_name=item_name,
                        quantity_on_hand=qty or 0,
                        reorder_point=reorder or 0,
                        unit_cost=price or 0.0,
                    )
                    new_products.append(wp)
                    existing_products[item_number] = wp
                    added += 1

            # Sync price -> ItemPrice so billing invoices use the same price
            if price is not None:
                ip = existing_prices.get(item_number)
                if ip:
                    ip.price = price
                    if item_name and not ip.item_name:
                        ip.item_name = item_name
                else:
                    ip = ItemPrice(item_number=item_number, item_name=item_name, price=price)
                    new_prices.append(ip)
                    existing_prices[item_number] = ip
                priced += 1

        if new_products:
            db.session.add_all(new_products)
        if new_prices:
            db.session.add_all(new_prices)
        db.session.commit()
        price_note = f', {len(price_from_sheet)} prices from PRICES sheet' if price_from_sheet else ''
        flash(
            f'Import complete: {added} products added, {updated} updated, {priced} prices synced{price_note}.',
            'success'
        )
        return redirect(url_for('billing.warehouse_inventory'))

    return render_template('billing_import.html')


# ---------- Metrics ----------
@billing_bp.route('/metrics')
@billing_required
def metrics():
    from sqlalchemy import func
    from datetime import timedelta

    # --- Trailer stats ---
    total_trailers = Trailer.query.count()
    completed_trailers = Trailer.query.filter_by(status='Completed').count()
    pending_trailers = Trailer.query.filter_by(status='Pending').count()
    in_progress_trailers = Trailer.query.filter_by(status='In Progress').count()

    flagged_counts = (
        db.session.query(
            InventoryResponse.item_name,
            InventoryResponse.item_number,
            func.count(InventoryResponse.id).label('count')
        )
        .filter(InventoryResponse.status.in_(['Missing', 'Red Tag']))
        .group_by(InventoryResponse.item_name, InventoryResponse.item_number)
        .order_by(func.count(InventoryResponse.id).desc())
        .limit(10)
        .all()
    )

    low_stock = WarehouseProduct.query.filter(
        WarehouseProduct.quantity_on_hand <= WarehouseProduct.reorder_point
    ).all()

    order_counts = (
        db.session.query(WarehouseOrder.status, func.count(WarehouseOrder.id))
        .group_by(WarehouseOrder.status).all()
    )
    order_status_map = {s: c for s, c in order_counts}

    # --- Monthly sales/purchases (last 12 months) ---
    twelve_ago = datetime.now() - timedelta(days=365)
    monthly_rows = (
        db.session.query(
            func.date_trunc('month', WarehouseOrder.created_at).label('month'),
            WarehouseOrder.order_type,
            func.count(WarehouseOrder.id).label('order_count'),
            func.coalesce(func.sum(WarehouseOrder.order_total), 0).label('total_value'),
        )
        .filter(WarehouseOrder.billed == True)
        .filter(WarehouseOrder.created_at >= twelve_ago)
        .group_by('month', WarehouseOrder.order_type)
        .order_by('month')
        .all()
    )
    # Pivot: {month_str -> {SALE: {count,value}, PURCHASE: {count,value}}}
    monthly_data = {}
    for row in monthly_rows:
        key = row.month.strftime('%b %Y') if row.month else 'Unknown'
        if key not in monthly_data:
            monthly_data[key] = {
                'SALE':     {'count': 0, 'value': 0.0},
                'PURCHASE': {'count': 0, 'value': 0.0},
            }
        otype = (row.order_type or 'SALE').upper()
        if otype in monthly_data[key]:
            monthly_data[key][otype] = {'count': row.order_count, 'value': float(row.total_value or 0)}
    months_list = list(monthly_data.keys())  # already ordered from query

    # --- Item date-range lookup ---
    item_search    = (request.args.get('item_search') or '').strip()
    from_date_str  = (request.args.get('from_date') or '').strip()
    to_date_str    = (request.args.get('to_date') or '').strip()
    otype_filter   = (request.args.get('order_type_filter') or '').strip().upper()
    item_results   = []
    item_totals    = {'qty': 0, 'value': 0.0}

    if item_search or (from_date_str and to_date_str):
        q = (
            db.session.query(
                WarehouseOrderLine.item_name,
                WarehouseOrderLine.item_number,
                WarehouseOrder.order_type,
                func.sum(WarehouseOrderLine.quantity).label('total_qty'),
                func.coalesce(func.sum(WarehouseOrderLine.line_total), 0).label('total_value'),
                func.count(WarehouseOrderLine.id).label('order_count'),
            )
            .join(WarehouseOrder, WarehouseOrderLine.order_id == WarehouseOrder.id)
            .filter(WarehouseOrder.billed == True)
        )
        if item_search:
            q = q.filter(
                db.or_(
                    WarehouseOrderLine.item_name.ilike(f'%{item_search}%'),
                    WarehouseOrderLine.item_number.ilike(f'%{item_search}%'),
                )
            )
        if from_date_str:
            try:
                q = q.filter(WarehouseOrder.created_at >= datetime.strptime(from_date_str, '%Y-%m-%d'))
            except ValueError:
                pass
        if to_date_str:
            try:
                end = datetime.strptime(to_date_str, '%Y-%m-%d') + timedelta(days=1)
                q = q.filter(WarehouseOrder.created_at < end)
            except ValueError:
                pass
        if otype_filter in ('SALE', 'PURCHASE'):
            q = q.filter(WarehouseOrder.order_type == otype_filter)

        item_results = (
            q.group_by(
                WarehouseOrderLine.item_name,
                WarehouseOrderLine.item_number,
                WarehouseOrder.order_type,
            )
            .order_by(func.sum(WarehouseOrderLine.quantity).desc())
            .all()
        )
        item_totals = {
            'qty':   sum(r.total_qty or 0 for r in item_results),
            'value': sum(float(r.total_value or 0) for r in item_results),
        }

    return render_template(
        'billing_metrics.html',
        total_trailers=total_trailers,
        completed_trailers=completed_trailers,
        pending_trailers=pending_trailers,
        in_progress_trailers=in_progress_trailers,
        flagged_counts=flagged_counts,
        low_stock=low_stock,
        order_status_map=order_status_map,
        monthly_data=monthly_data,
        months_list=months_list,
        item_results=item_results,
        item_totals=item_totals,
        item_search=item_search,
        from_date_str=from_date_str,
        to_date_str=to_date_str,
        otype_filter=otype_filter,
    )


# ---------- Specialty Tools ----------
@billing_bp.route('/specialty-tools')
@billing_required
def specialty_tools():
    q = (request.args.get('q') or '').strip().lower()
    tools = SpecialtyTool.query.order_by(SpecialtyTool.item_name).all()
    if q:
        tools = [t for t in tools if q in (t.item_name or '').lower() or q in (t.item_number or '').lower()]
    return render_template('billing_specialty_tools.html', tools=tools, q=q)


@billing_bp.route('/specialty-tools/add', methods=['POST'])
@billing_required
def specialty_tool_add():
    item_number = (request.form.get('item_number') or '').strip()
    item_name = (request.form.get('item_name') or '').strip()
    if not item_number or not item_name:
        flash('Item number and name are required.', 'danger')
        return redirect(url_for('billing.specialty_tools'))
    if SpecialtyTool.query.filter_by(item_number=item_number).first():
        flash('A tool with that item number already exists.', 'warning')
        return redirect(url_for('billing.specialty_tools'))
    try:
        price = float(request.form.get('price') or 0)
    except ValueError:
        price = 0.0
    try:
        quantity = int(request.form.get('quantity') or 0)
    except ValueError:
        quantity = 0
    db.session.add(SpecialtyTool(item_number=item_number, item_name=item_name,
                                  price=price, quantity=quantity))
    db.session.commit()
    flash('Specialty tool added.', 'success')
    return redirect(url_for('billing.specialty_tools'))


@billing_bp.route('/specialty-tools/<int:tool_id>/edit', methods=['POST'])
@billing_required
def specialty_tool_edit(tool_id):
    tool = SpecialtyTool.query.get_or_404(tool_id)
    tool.item_number = (request.form.get('item_number') or tool.item_number).strip()
    tool.item_name = (request.form.get('item_name') or tool.item_name).strip()
    try:
        tool.price = float(request.form.get('price') or 0)
    except ValueError:
        pass
    try:
        tool.quantity = int(request.form.get('quantity') or 0)
    except ValueError:
        pass
    db.session.commit()
    flash('Tool updated.', 'success')
    return redirect(url_for('billing.specialty_tools'))


@billing_bp.route('/specialty-tools/<int:tool_id>/delete', methods=['POST'])
@billing_required
def specialty_tool_delete(tool_id):
    tool = SpecialtyTool.query.get_or_404(tool_id)
    db.session.delete(tool)
    db.session.commit()
    flash('Tool deleted.', 'info')
    return redirect(url_for('billing.specialty_tools'))
